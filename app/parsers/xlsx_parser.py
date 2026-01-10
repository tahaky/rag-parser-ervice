from typing import Dict, Any, List
from openpyxl import load_workbook

from app.parsers.base import BaseParser


class XlsxParser(BaseParser):
    """Parser for XLSX spreadsheets."""

    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse XLSX spreadsheet into structured JSON."""
        wb = load_workbook(file_path, data_only=True, read_only=True)

        # Extract metadata
        props = wb.properties
        metadata = self.build_metadata(
            title=props.title or "",
            creator=props.creator or "",
            subject=props.subject or "",
            created=str(props.created) if props.created else None,
            modified=str(props.modified) if props.modified else None,
        )

        # Parse sheets
        sheets = []
        total_text_length = 0
        total_tables = 0
        total_cells = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            sheet_data = {
                "sheet_name": sheet_name,
                "rows": [],
                "row_count": 0,
                "col_count": 0,
            }

            # Extract data
            rows = []
            max_col = 0
            
            for row in ws.iter_rows(values_only=True):
                # Convert row to list and filter out completely empty rows
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell for cell in row_data):  # Only add if row has content
                    rows.append(row_data)
                    max_col = max(max_col, len(row_data))
                    total_cells += len([c for c in row_data if c])
                    
                    # Add to text length
                    for cell in row_data:
                        if cell:
                            total_text_length += len(cell)

            sheet_data["rows"] = rows
            sheet_data["row_count"] = len(rows)
            sheet_data["col_count"] = max_col

            # Calculate hash for the sheet
            if rows:
                sheet_text = " ".join(" ".join(row) for row in rows)
                sheet_data["hash"] = self.calculate_hash(sheet_text)
                total_tables += 1  # Count each sheet as a table

            sheets.append(sheet_data)

        wb.close()

        # Build structure
        structure = {
            "format": "xlsx",
            "sheets": sheets,
        }

        # Build stats
        stats = self.build_stats(
            total_pages=len(sheets),
            total_text_length=total_text_length,
            total_tables=total_tables,
            total_images=0,
            total_sheets=len(sheets),
            total_cells=total_cells,
        )

        return {
            "format": "xlsx",
            "metadata": metadata,
            "structure": structure,
            "stats": stats,
        }
